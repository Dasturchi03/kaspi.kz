import os
import sys
import json
import asyncio
import openpyxl
from openpyxl.utils import get_column_letter
from aiogram import Bot, Dispatcher, executor
from aiogram.types import Message, InputFile, Document, File
from pyppeteer import launch
from bs4 import BeautifulSoup
from apscheduler.schedulers.asyncio import AsyncIOScheduler


with open('settings.json', 'r') as settings_file:
    dc = json.load(settings_file)
    INTERVAL = int(dc['interval'])
    bot_token = dc['bot_token']


bot = Bot(bot_token)
dp = Dispatcher(bot)


@dp.message_handler(commands=['urls'])
async def get_urls_file(message: Message):
    with open('settings.json', 'r') as file:
        dc = json.load(file)
        MESSAGE = await bot.send_document(message.chat.id, InputFile('files/urls.txt'))
        dc['users'][str(message.chat.id)] = MESSAGE.message_id
    with open('settings.json', 'w') as file:
        file.write(json.dumps(dc, ensure_ascii=False, indent=2))


@dp.message_handler(commands=['getfile'])
async def get_xlsx_file(message: Message):
    await write_xlsx()
    with open('settings.json', 'r') as file:
        dc = json.load(file)
        MESSAGE = await bot.send_document(message.chat.id, InputFile('files/DATA.xlsx'))
        dc['users'][str(message.chat.id)] = MESSAGE.message_id
    with open('settings.json', 'w') as file:
        file.write(json.dumps(dc, ensure_ascii=False, indent=2))
    if not STARTED:
        await main()

@dp.message_handler(content_types=['document'])
async def add_urls_file(message: Message):
    with open('settings.json', 'r') as file:
        dc = json.load(file)
        file_id = message.document.file_id
        file_info = await bot.get_file(file_id)
        lines = await download_file(file_info)
        for i in lines:
            print(i)
        MESSAGE = await message.reply("Принял")
        dc['users'][str(message.chat.id)] = MESSAGE.message_id
    with open('settings.json', 'w') as file:
        file.write(json.dumps(dc, ensure_ascii=False, indent=2))
    if not STARTED:
        await main()

@dp.message_handler()
async def start_bot(message: Message, state):
    with open('settings.json', 'r') as file:
        dc = json.load(file)
        mess = "Используйте команду /getfile, чтобы получить проанализированный файл Excel.\n\n"
        mess += "Используйте команду /urls, чтобы получить файл, содержащий URL-адреса.\n\n"
        mess += "Чтобы сбросить URL-адреса, отправьте мне файл .txt с URL-адресами."
        MESSAGE = await bot.send_message(message.chat.id, mess)
        dc['users'][str(message.chat.id)] = MESSAGE.message_id
    with open('settings.json', 'w') as file:
        file.write(json.dumps(dc, ensure_ascii=False, indent=2))
    if not STARTED:
        await main()


async def download_file(file_info: File):
    file_path = file_info.file_path
    try:
        os.mkdir('files')
    except:
        pass
    await bot.download_file(file_path, 'files/urls.txt')
    with open('files/urls.txt', 'r') as file:
        return file.readlines()


async def get_urls():
    data = []
    with open('files/urls.txt', 'r') as file:
        for i in file.readlines():
            data.append(i.strip('\n').strip())
    return data


async def main():
    global STARTED
    STARTED = True
    try:
        with open('settings.json', 'r') as file:
            dc = json.load(file)
            for i, j in dc['users'].items():
                await bot.edit_message_text("Загрузка данных...", i, j)
    except:
        pass
    sys.stdout.write('\x1b[1A')
    sys.stdout.write("\x1b[2K")
    print("Обновление информации...")
    urls = await get_urls()
    brow = await launch(headless=True, options={'args': ['--no-sandbox']})
    page = await brow.newPage()
    cookies = [{'name': 'layout', 'value': 'd', 'domain': 'kaspi.kz', 'path': '/shop', 'expires': 1726644470.633389, 'size': 7, 'httpOnly': False, 'secure': False, 'session': False}, {'name': 'k_stat', 'value': '74791a40-1e09-45f8-a9ed-d4ec9e7b7359', 'domain': 'kaspi.kz', 'path': '/', 'expires': 1726644470.844751, 'size': 42, 'httpOnly': False, 'secure': False, 'session': False}, {'name': 'ks.tg', 'value': '84', 'domain': 'kaspi.kz', 'path': '/', 'expires': 1726644470.928354, 'size': 7, 'httpOnly': False, 'secure': False, 'session': False}, {'name': 'kaspi.storefront.cookie.city', 'value': '632210000', 'domain': 'kaspi.kz', 'path': '/', 'expires': 1726644474, 'size': 37, 'httpOnly': False, 'secure': False, 'session': False}]
    await page.setCookie(*cookies)
    await reset_json()
    for url in urls:
        # try:
            await page.goto(url)
            await page.waitFor(200)
            cont = await page.content()
            await get_data(cont, url)
        # except Exception as err:
            # print(err)
    await write_xlsx()
    
    await brow.close()
    sys.stdout.write('\x1b[1A')
    sys.stdout.write("\x1b[2K")
    print("Информация обновлена")
    try:
        with open('settings.json', 'r') as file:
            dc = json.load(file)
            for i, j in dc['users'].items():
                await bot.edit_message_text("Данные загружены✅", i, j)
    except:
        pass


async def get_data(text, url):
    soup = BeautifulSoup(text, 'lxml')
    div = soup.find('div', id='ItemView')
    h1 = soup.find('h1', class_='item__heading')
    name = h1.text.strip()
    d1 = soup.find('div', class_='item__price-once')
    price = d1.text.strip()
    with open('files/data.json', 'r', encoding='UTF-8') as file:
        dc = json.load(file)
    dc[url] = [name, price]
    
    with open('files/data.json', 'w', encoding='utf-8') as file:
        file.write(json.dumps(dc, indent=2, ensure_ascii=False))


async def reset_json():
    with open('files/data.json', 'w', encoding='utf-8') as file:
        file.write(json.dumps({}))


async def write_xlsx():
    try:
        wk = openpyxl.load_workbook('files/DATA.xlsx')
    except:
        wk = openpyxl.Workbook()
    sheet = wk.active
    wk.remove(sheet)
    wk.create_sheet("Sheet")
    sheet = wk['Sheet']
    with open('files/data.json', 'r', encoding='UTF-8') as file:
        dc = json.load(file)
    for i, j in dc.items():
        sheet.append([i] + j)
    sheet.column_dimensions[get_column_letter(1)].width = 90
    sheet.column_dimensions[get_column_letter(2)].width = 50
    try:
        wk.save('files/DATA.xlsx')
    except:
        pass


if __name__ == '__main__':
    sched = AsyncIOScheduler()
    print(f"Интервал установлен на {INTERVAL} секунд.")
    print("Вы можете изменить это в файле settings.json.")
    sched.add_job(main, 'interval', seconds=INTERVAL)
    STARTED = False
    sched.start()
    print('Бот запущен')
    executor.start_polling(dp)
